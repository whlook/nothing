#!/usr/local/anaconda3/bin/python
import sys
import time
import datetime
import openpyxl as opx
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

def read_excel(excel_file_path = "paiban.xlsx", query_days = 7):
    wb = opx.load_workbook(excel_file_path)
    sheetnames = wb.sheetnames
    month = datetime.date.today().month
    monthstr = str(month)+"月"
    sheet = wb[sheetnames[len(sheetnames)-1]]

    for sn in sheetnames:
        if monthstr in sn:
            sheet = wb[sn]
            break

    today = datetime.date.today().day
    res = {}
    start_col = 1

    #first: positioning name location and date start location
    for a in range(3,40):
        cell = sheet.cell(2,a)
        day = cell.value

        if day >= today:
            cell = sheet.cell(3,a)
            week = cell.value
            if week != u"一":
                continue
            start_col = a
            break

    #second: make list & dict
    for row in range(4,50):
        cell = sheet.cell(row,2)
        name = cell.value
        schedul = 1 #schedul: 1-班　2-休
        sch_list = {}

        if name is None:
            continue
        days = 0
        for col in range(start_col,40):
            days += 1
            if days > query_days:
                break
            cell = sheet.cell(row,col)
            schedul_val = cell.value
            if schedul_val is None:
                break
            if schedul_val == u"班":
                schedul = 1
            elif schedul_val == u"休" or schedul_val == u"调休":
                schedul = 2
            cell = sheet.cell(2,col)
            sch_list[int(cell.value)] = schedul

        res[name] = sch_list

    return res

def online_operate(staffs = {},oa_name = "username",oa_pass = "passwd"):
    # {"xxx":{27:1,28:1,29:2}}
    if staffs:
        today = datetime.date.today()
        prename = "00"
        chrome = webdriver.Chrome()
        time.sleep(2)
        chrome.get("https://oa.xxxx.com")
        time.sleep(2)
        ele=chrome.find_element_by_id("username")
        ele.clear()
        ele.send_keys(oa_name)
        ele=chrome.find_element_by_id("userpass")
        ele.clear()
        ele.send_keys(oa_pass)
        btn=chrome.find_element_by_id("Button1")
        btn.click()
        time.sleep(2)
        oa_sys=chrome.find_element_by_class_name("list-item")
        oa_sys.click()
        time.sleep(2)
        handles=chrome.window_handles
        chrome.switch_to_window(handles[1])
        time.sleep(2)
        backend=chrome.find_element_by_link_text(u"后台管理")
        backend.click()
        time.sleep(2)
        kaoqinpeizhi=chrome.find_element_by_id("li_First_10")
        kaoqinpeizhi.click()
        time.sleep(2)
        paibanpeizhi=chrome.find_element_by_link_text(u"排班配置")
        paibanpeizhi.click()
        time.sleep(1)
        for name in staffs:
            staff_name=chrome.find_element_by_id("USER_NAME")
            staff_name.click()
            staff_name.clear()
            staff_name.send_keys(name)
            query=chrome.find_element_by_id("btn_search")
            query.click()
            time.sleep(2)
            chrome.switch_to_default_content()
            try:
                setup=chrome.find_element_by_link_text(u'[调整]')
                setup.click()
            except:
                print(name + " not found")
                continue
            time.sleep(1)
            chrome.switch_to.frame("OpenAdjustSetting")

            for day in staffs[name]:
               ele=chrome.find_element_by_xpath("//td")
               strs = "//span[@arrangedate='"+str(today.year)+'-'+prename[0:2-len(str(today.month))]+str(today.month)+'-'+prename[0:2-len(str(day))]+str(day)+"']"
               mode = staffs[name][day]
               date=ele.find_element_by_xpath(strs)
               date.click()
               time.sleep(0.5)
               select = chrome.find_element_by_xpath("//option[@value='135']")
               select.click()
               select = chrome.find_element_by_xpath("//option[@value='181']")
               select.click()
               if mode == 1:
                   select = chrome.find_element_by_xpath("//option[@value='1']")
                   select.click()
               elif mode == 2:
                   select = chrome.find_element_by_xpath("//option[@value='2']")
                   select.click()

               save = chrome.find_element_by_id("addmanager")
               save.click()
               time.sleep(1)
               save.send_keys(Keys.ESCAPE)
               time.sleep(1)

            select = chrome.find_element_by_xpath("//input[@value='关闭']")
            select.click()
            time.sleep(1)
            print(">> "+name+" ok <<")

if __name__ == '__main__':
    args = sys.argv
    a = read_excel(args[1],7)
    print(a)
    online_operate(staffs=a)
